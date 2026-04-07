from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Iterable, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .analysis import compute_matrix

MATRIX_EXPORT_HEADERS = [
    "治疗领域",
    "疾病种类",
    "药品中文名",
    "药品英文名",
    "nctId",
    "原研机构",
    "所有研究机构",
    "靶点组合",
    "最高阶段",
    "最高阶段日期",
    "最高阶段分值",
    "是否联用",
]
PIPELINE_EXPORT_HEADERS = [
    "治疗领域",
    "疾病种类",
    "靶点",
    "药品中文名",
    "药品英文名",
    "nctId",
    "原研机构",
    "所有研究机构",
    "研发阶段",
    "阶段分值",
    "最高阶段日期",
]

MATRIX_EXPORT_SHEET_NAME = "靶点组合竞争矩阵"
PIPELINE_EXPORT_SHEET_NAME = "靶点研发进展格局"
EXPORT_HEADER_FILL = PatternFill("solid", fgColor="DCE7FF")
EXPORT_HEADER_FONT = Font(bold=True, color="1F3763")
EXPORT_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
EXPORT_BODY_ALIGNMENT = Alignment(vertical="center")
MATRIX_EXPORT_COLUMN_WIDTHS = [16, 24, 18, 20, 18, 18, 26, 22, 14, 14, 12, 10]
PIPELINE_EXPORT_COLUMN_WIDTHS = [16, 24, 18, 18, 20, 18, 18, 26, 14, 12, 14]


def _format_targets(targets: Sequence[str]) -> str:
    return " + ".join(targets)


def _build_matrix_export_row(record: dict, targets: Sequence[str]) -> List[object]:
    return [
        record["ta"],
        record["disease"],
        record["drug_name_cn"],
        record["drug_name_en"],
        record["nct_id"],
        record["originator"],
        record["research_institute"],
        _format_targets(targets),
        record["stage_value"],
        record["highest_trial_date"],
        record["score"],
        "是" if record["is_combo"] else "否",
    ]


def _build_pipeline_export_row(record: dict, target: str) -> List[object]:
    return [
        record["ta"],
        record["disease"],
        target,
        record["drug_name_cn"],
        record["drug_name_en"],
        record["nct_id"],
        record["originator"],
        record["research_institute"],
        record["stage_value"],
        record["score"],
        record["highest_trial_date"],
    ]


def _sort_records(records: Iterable[dict]) -> List[dict]:
    return sorted(
        records,
        key=lambda item: (
            -item["score"],
            -(int((item["highest_trial_date"] or "0000-00-00").replace("-", ""))),
            item["drug_name_en"] or item["drug_name_cn"] or "",
            item["nct_id"] or "",
            item["disease"] or "",
        ),
    )


def build_matrix_export_rows(
    records: Sequence[dict],
    selected_targets: Optional[Sequence[str]] = None,
    hide_no_combo: bool = False,
) -> List[List[object]]:
    matrix_data = compute_matrix(
        records,
        selected_targets=selected_targets,
        hide_no_combo=hide_no_combo,
    )
    display_targets = set(matrix_data["targets"])
    if not display_targets:
        return []

    rows: List[List[object]] = []
    for record in _sort_records(records):
        matched_targets = [target for target in record["target_list"] if target in display_targets]
        if not matched_targets:
            continue
        rows.append(_build_matrix_export_row(record, matched_targets))
    return rows


def build_pipeline_export_rows(
    records: Sequence[dict],
    selected_targets: Optional[Sequence[str]] = None,
    include_combo: bool = True,
) -> List[List[object]]:
    target_filter = None if not selected_targets else set(selected_targets)
    rows: List[List[object]] = []

    for record in _sort_records(records):
        if not include_combo and len(record["target_list"]) > 1:
            continue

        matched_targets = record["target_list"]
        if target_filter is not None:
            matched_targets = [target for target in matched_targets if target in target_filter]
            if not matched_targets:
                continue

        for target in matched_targets:
            rows.append(_build_pipeline_export_row(record, target))

    return rows


def build_excel_bytes(
    rows: Sequence[Sequence[object]],
    headers: Sequence[object],
    sheet_name: str,
    column_widths: Sequence[int],
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.freeze_panes = "A2"
    worksheet.append(list(headers))

    for row in rows:
        worksheet.append(list(row))

    for cell in worksheet[1]:
        cell.fill = EXPORT_HEADER_FILL
        cell.font = EXPORT_HEADER_FONT
        cell.alignment = EXPORT_HEADER_ALIGNMENT

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = EXPORT_BODY_ALIGNMENT

    for index, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width

    worksheet.auto_filter.ref = worksheet.dimensions

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_export_filename(export_type: str) -> str:
    today = datetime.now().strftime("%Y-%m-%d")
    return f"Apex_Target_Intelligence_{export_type}_{today}.xlsx"
